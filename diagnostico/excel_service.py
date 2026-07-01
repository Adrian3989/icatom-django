from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

def generar_excel_historial(diagnosticos, usuario):
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    # Estilos
    header_fill = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezado
    ws['A1'] = "ICATOM S.A. — Historial de Diagnósticos"
    ws['A1'].font = Font(bold=True, size=14, color="1b5e20")
    ws.merge_cells('A1:H1')

    ws['A2'] = f"Usuario: {usuario.nombre}"
    ws['A3'] = f"Total de diagnósticos: {diagnosticos.count()}"

    # Encabezados de tabla
    headers = ['#', 'Fecha', 'Enfermedad', 'Severidad', 'Sector', 'Síntomas', 'Tratamiento', 'Usuario']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    for row_idx, diag in enumerate(diagnosticos, 6):
        ws.cell(row=row_idx, column=1).value = diag.id
        ws.cell(row=row_idx, column=2).value = diag.fecha_consulta.strftime('%d/%m/%Y %H:%M')
        ws.cell(row=row_idx, column=3).value = diag.enfermedad.nombre if diag.enfermedad else '—'
        ws.cell(row=row_idx, column=4).value = diag.severidad.capitalize() if diag.severidad else '—'
        ws.cell(row=row_idx, column=5).value = diag.sector.nombre if diag.sector else '—'
        ws.cell(row=row_idx, column=6).value = diag.sintomas if diag.sintomas else '—'
        ws.cell(row=row_idx, column=7).value = diag.tratamiento if diag.tratamiento else '—'
        ws.cell(row=row_idx, column=8).value = diag.usuario.nombre

        # Bordes
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    # Ancho de columnas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 15

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer